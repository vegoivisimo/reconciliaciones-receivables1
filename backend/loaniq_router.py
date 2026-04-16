from fastapi import APIRouter, File, UploadFile, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from pipeline import load_santix, load_loaniq, build_groups
from matching import reconcile_all
from santix_loaniq_agent import chat_with_agent, generate_ai_summary
from schemas import ChatRequest, OverrideLogEntry
from database import init_db, save_resolution, log_override, get_override_log

router = APIRouter(prefix="/api/loaniq", tags=["loaniq"])

init_db()


@router.post("/reconcile")
async def reconcile(
    santix: UploadFile = File(...),
    loaniq: UploadFile = File(...),
):
    try:
        santix_bytes = await santix.read()
        loaniq_bytes = await loaniq.read()
        santix_df = load_santix(io.BytesIO(santix_bytes))
        loaniq_df = load_loaniq(io.BytesIO(loaniq_bytes))
        groups = build_groups(santix_df)
        result = reconcile_all(groups, loaniq_df)
        return result
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivos: {str(e)}")


@router.post("/chat")
async def chat(req: ChatRequest):
    try:
        reply = chat_with_agent(
            [m.dict() for m in req.messages],
            req.group_context,
        )
        return {"reply": reply}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class OverrideRequest(BaseModel):
    group_key: str
    loaniq_alias: str
    santix_debtor: str = ""
    sum_paid: float = 0.0
    original_tier: str = "UNKNOWN"
    operator: str = "operator"


@router.post("/override")
async def override_match(req: OverrideRequest):
    save_resolution(
        group_key=req.group_key,
        loaniq_alias=req.loaniq_alias,
        santix_debtor=req.santix_debtor,
        sum_paid=req.sum_paid,
        resolved_by=req.operator,
    )
    log_override(
        group_key=req.group_key,
        santix_debtor=req.santix_debtor,
        original_tier=req.original_tier,
        loaniq_alias=req.loaniq_alias,
        operator=req.operator,
    )
    return {
        "group_key":    req.group_key,
        "loaniq_alias": req.loaniq_alias,
        "status":       "MANUAL_OVERRIDE",
    }


@router.get("/overrides", response_model=list[OverrideLogEntry])
async def get_overrides():
    return get_override_log()


@router.post("/ai-summary")
async def ai_summary(request: Request):
    try:
        body    = await request.json()
        summary = body.get("summary", {})
        groups  = body.get("groups", [])
        text    = generate_ai_summary(summary, groups)
        return {"summary": text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Colores por tier ──────────────────────────────────────────────────────────
TIER_FILL = {
    "MATCH_EXACT":            PatternFill("solid", fgColor="C6EFCE"),   # verde
    "MATCH_CROSS_PREFIX":     PatternFill("solid", fgColor="D9D2E9"),   # lila
    "REVIEW_OVERFLOW":        PatternFill("solid", fgColor="FFEB9C"),   # amarillo
    "REVIEW_DEBTOR_MISMATCH": PatternFill("solid", fgColor="FCE4D6"),   # salmon
    "SUGGESTED":              PatternFill("solid", fgColor="FFF2CC"),   # amarillo palido
    "NO_MATCH":               PatternFill("solid", fgColor="FFDDC1"),   # naranja palido
    "LEARNED_MATCH":          PatternFill("solid", fgColor="E8D5F0"),   # rosa lila
    "MANUAL_OVERRIDE":        PatternFill("solid", fgColor="EAD1DC"),   # rosa
}

TIER_LEGEND = {
    "MATCH_EXACT":            "Alias exacto + deudor validado — auto-reconciliado",
    "MATCH_CROSS_PREFIX":     "Cross-prefix: deudor + OA coinciden — auto-reconciliado",
    "REVIEW_OVERFLOW":        "Alias OK pero importe pagado supera Original Amount — revisar",
    "REVIEW_DEBTOR_MISMATCH": "Alias exacto pero deudor no coincide — revisar asignación",
    "SUGGESTED":              "Múltiples candidatos ambiguos — elección manual requerida",
    "NO_MATCH":               "Sin candidatos — investigación manual",
    "LEARNED_MATCH":          "Match de override previo re-validado",
    "MANUAL_OVERRIDE":        "Override manual del operador",
}

NO_FILL     = PatternFill(fill_type=None)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
YELLOW      = PatternFill("solid", fgColor="FFFF00")


@router.post("/export-updated")
async def export_updated_loaniq(
    santix: UploadFile = File(...),
    loaniq: UploadFile = File(...),
):
    try:
        santix_bytes = await santix.read()
        loaniq_bytes = await loaniq.read()
        santix_df  = load_santix(io.BytesIO(santix_bytes))
        loaniq_df  = load_loaniq(io.BytesIO(loaniq_bytes))
        groups     = build_groups(santix_df)
        result     = reconcile_all(groups, loaniq_df)

        loaniq_updated = result.get("loaniq_updated", [])
        summary        = result.get("summary", {})

        wb = openpyxl.Workbook()

        # ── Hoja 1: LOANIQ Reconciliado ────────────────────────────────────
        ws1 = wb.active
        ws1.title = "LOANIQ Reconciliado"

        if loaniq_updated:
            headers  = [k for k in loaniq_updated[0].keys() if k != "_tier"]
            curr_col = headers.index("Current Amount") + 1 if "Current Amount" in headers else None

            for col_idx, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

            for row_idx, rec in enumerate(loaniq_updated, 2):
                tier       = rec.get("_tier", "UNMATCHED")
                is_matched = tier not in ("UNMATCHED", "REVIEW_OVERFLOW", "REVIEW_DEBTOR_MISMATCH",
                                          "SUGGESTED", "NO_MATCH")
                for col_idx, h in enumerate(headers, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=rec.get(h))
                    if col_idx == curr_col and is_matched:
                        cell.fill = YELLOW

            for col_idx, h in enumerate(headers, 1):
                ws1.column_dimensions[get_column_letter(col_idx)].width = max(len(str(h)) + 4, 14)

        # ── Hoja 2: Detalle Motor ───────────────────────────────────────────
        ws2 = wb.create_sheet("Detalle Motor")

        if loaniq_updated:
            detail_headers = headers + ["Tier"]
            for col_idx, h in enumerate(detail_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

            for row_idx, rec in enumerate(loaniq_updated, 2):
                tier = rec.get("_tier", "UNMATCHED")
                fill = TIER_FILL.get(tier, NO_FILL)
                for col_idx, h in enumerate(headers, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=rec.get(h))
                    cell.fill = fill
                tier_cell = ws2.cell(row=row_idx, column=len(headers) + 1, value=tier)
                tier_cell.fill = fill

            for col_idx in range(1, len(detail_headers) + 1):
                ws2.column_dimensions[get_column_letter(col_idx)].width = 16

            legend_row = len(loaniq_updated) + 4
            ws2.cell(row=legend_row, column=1, value="LEYENDA").font = Font(bold=True, size=12)
            for i, (tier_name, desc) in enumerate(TIER_LEGEND.items(), 1):
                r = legend_row + i
                cell = ws2.cell(row=r, column=1, value=tier_name)
                cell.fill = TIER_FILL.get(tier_name, NO_FILL)
                ws2.cell(row=r, column=2, value=desc)

        # ── Hoja 3: Resumen Reconciliación ─────────────────────────────────
        ws3 = wb.create_sheet("Resumen Reconciliación")

        for col_idx, h in enumerate(["Métrica", "Valor"], 1):
            cell = ws3.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        metrics = [
            ("SANTIX Grupos",            summary.get("santix_groups", 0)),
            ("SANTIX Facturas",           summary.get("santix_invoices", 0)),
            ("LOANIQ Filas",              summary.get("loaniq_rows", 0)),
            ("", ""),
            ("✓ MATCH_EXACT",             summary.get("match_exact", 0)),
            ("✓ MATCH_CROSS_PREFIX",      summary.get("match_cross_prefix", 0)),
            ("✓ LEARNED_MATCH",           summary.get("learned_match", 0)),
            ("", ""),
            ("⚠ REVIEW_OVERFLOW",         summary.get("review_overflow", 0)),
            ("⚠ REVIEW_DEBTOR_MISMATCH",  summary.get("review_debtor_mismatch", 0)),
            ("⚠ SUGGESTED",               summary.get("suggested", 0)),
            ("✗ NO_MATCH",                summary.get("no_match", 0)),
            ("", ""),
            ("STP Rate %",               summary.get("stp_rate_pct", 0)),
            ("", ""),
            ("Total SANTIX EUR",          summary.get("total_santix_eur", 0)),
            ("Total LOANIQ EUR",          summary.get("total_loaniq_eur", 0)),
        ]

        for row_idx, (metric, value) in enumerate(metrics, 2):
            ws3.cell(row=row_idx, column=1, value=metric)
            cell_v = ws3.cell(row=row_idx, column=2, value=value)
            if isinstance(value, float) and value > 1000:
                cell_v.number_format = '#,##0.00'

        ws3.column_dimensions["A"].width = 32
        ws3.column_dimensions["B"].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"LOANIQ_reconciliado_{date.today().isoformat()}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
